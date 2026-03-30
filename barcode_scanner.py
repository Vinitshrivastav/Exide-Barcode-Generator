"""
Exide Barcode Scanner → Excel
2026-C-A-1-01045 scan karo → 2026CA101045 Excel mein save hoga
"""
import sys, os, msvcrt
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    print("  openpyxl install ho raha hai...")
    subprocess.run([sys.executable,"-m","pip","install","openpyxl"],check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE   = Path(__file__).parent
EXCEL  = BASE / "Exide_Scanned_Serials.xlsx"

def get_wb():
    if EXCEL.exists():
        wb=openpyxl.load_workbook(str(EXCEL)); ws=wb.active
        existing={str(ws.cell(r,2).value).strip() for r in range(2,ws.max_row+1) if ws.cell(r,2).value}
        return wb,ws,existing

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Scanned Serials"
    hf=PatternFill("solid",start_color="1D6F42")
    hfont=Font(bold=True,color="FFFFFF",size=12,name="Arial")
    ha=Alignment(horizontal="center",vertical="center")
    ht=Side(style="thin",color="999999")
    hb=Border(left=ht,right=ht,top=ht,bottom=ht)
    for col,hdr in enumerate(["S.No.","Serial Number","Date","Time"],1):
        c=ws.cell(1,col,hdr); c.font=hfont; c.fill=hf; c.alignment=ha; c.border=hb
    ws.row_dimensions[1].height=26
    ws.column_dimensions["A"].width=8
    ws.column_dimensions["B"].width=22
    ws.column_dimensions["C"].width=16
    ws.column_dimensions["D"].width=12
    wb.save(str(EXCEL))
    return wb,ws,set()

def save_scan(raw):
    # Store as-is (with hyphens) e.g. 2026-C-A-1-01045
    cleaned = raw.strip()
    if not cleaned: return "empty"
    wb,ws,existing=get_wb()
    if cleaned in existing: return "duplicate"

    thin=Side(style="thin",color="DDDDDD")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ali=Alignment(horizontal="center",vertical="center")
    font=Font(name="Courier New",size=11)
    alt=PatternFill("solid",start_color="F0FAF4")
    nr=ws.max_row+1; now=datetime.now()
    for col,val in enumerate([nr-1,cleaned,now.strftime("%d/%m/%Y"),now.strftime("%H:%M:%S")],1):
        c=ws.cell(nr,col,val); c.font=font; c.alignment=ali; c.border=bdr
        if nr%2==0: c.fill=alt
    wb.save(str(EXCEL))
    return "saved"

def main():
    os.system("cls")
    print("="*52)
    print("   EXIDE SERIAL SCANNER  →  EXCEL SAVER")
    print("="*52)
    print(f"   File: {EXCEL.name}")
    print(f"   Folder: {BASE}")
    print()
    print("   ✔  Scan karo → Excel mein save")
    print("   ✔  2026-C-A-1-01045 → 2026CA101045")
    print("   ✔  Duplicate → skip")
    print()
    print("   Band karne ke liye: window band karo")
    print("="*52)
    print("\n   Ready — scan karo...\n")

    buf=""; total=0
    while True:
        try:
            ch=msvcrt.getwch()
            if ch in('\r','\n'):
                raw=buf.strip(); buf=""
                if not raw: continue
                result=save_scan(raw)
                ts=datetime.now().strftime("%H:%M:%S")
                if result=="saved":
                    total+=1
                    cleaned=raw.replace("-","")
                    print(f"   [{ts}]  ✓ SAVED  →  {cleaned}  (Total: {total})")
                elif result=="duplicate":
                    print(f"   [{ts}]  ⚠ DUPLICATE  →  {cleaned}  (skip)")
            elif ch=='\x03': raise KeyboardInterrupt
            else: buf+=ch
        except KeyboardInterrupt:
            print(f"\n   Band ho gaya. Total: {total}")
            input("   Enter dabao..."); break
        except Exception: buf=""

if __name__=="__main__":
    main()
